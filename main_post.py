<<<<<<< HEAD
# -*- coding: utf-8 -*-
"""
Created on Wed Apr 27 16:50:16 2016

@author: nas-263
"""

# start program ------------------------------------------------------------

import adodbapi
import sys
from PyQt4 import QtCore, QtGui
from ui import post_text
import os
import openpyxl


class StartQT4(QtGui.QMainWindow):
    def __init__(self, parent=None):
        QtGui.QWidget.__init__(self, parent)
        self.ui = post_text.Ui_MainWindow()
        self.ui.setupUi(self)

        QtCore.QObject.connect(self.ui.pushButton, QtCore.SIGNAL("clicked()"), self.updateUi)
        #QtCore.QObject.connect(self.ui.pushButton_print, QtCore.SIGNAL("clicked()"), self.print_data)
        QtCore.QObject.connect(self.ui.pushButton_print, QtCore.SIGNAL("clicked()"), self.error)


    def error(self):
        self.ui.textEdit.append("現在、使用できません。")

    def updateUi(self):
        # input post No. or addressa
        global text
        text = (str(self.ui.lineEdit.text()))        
        self.ui.textEdit.setText("")

        # connect to the database
        database = r"./data/Database1.mdb"
        constr = 'Provider=Microsoft.Jet.OLEDB.4.0; Data Source=%s' %database
        conn = adodbapi.connect(constr)
        
        # create a cursor
        cur = conn.cursor()
        
        # extract all the data
        tablename = "24mie"
        sql = r"select 郵便番号, 県名, 市町村名, 住所 from %s" %tablename
        #sql = "select * from %s" %tablename
        cur.execute(sql)
        
        # show the cursor
        row2 = list()
        for row in cur:
            row1 = row.郵便番号[:3]+"-"+row.郵便番号[-4:]+"："+row.県名+row.市町村名+row.住所
            if row1.find(text) >= 0:
                row2 += row1,
        row2.sort()

        # close the cursor and connection
        cur.close()
        conn.close()
    
        # set data
        global post
        global address
        global count
        post = list()
        address = list()
        count = 0
        for i in row2:
            self.ui.textEdit.append(i)
            post    += i[:8],
            address += i[9:],
            count   += 1

    def print_data(self):
        wb = openpyxl.load_workbook('workfile.xlsx')
        ws = wb.get_active_sheet()
        i = 0
        alpha = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        for column in ws.columns:
            columns_count = ws.columns.index(column)
            columns_count = alpha[columns_count]
                        
            for cell in column:
                rows_count = column.index(cell) + 1

                if cell.value == "{d}":
                    rows_i = rows_count
                if cell.value == "{text}":
                    rows_x = rows_count
                    columns_x = columns_count
                if cell.value == "{post}":
                    rows_y = rows_count
                    columns_y = columns_count
                if cell.value == "{address}":
                    columns_z = columns_count
                if cell.value == "{page}":
                    rows_p = rows_count
                    columns_p = columns_count

        # write the data
        a = 0
        i = rows_i - rows_y + 1
        i2 = int((count - 1) / i) + 1

        for p in range(i2):
            page = p + 1
            rows_z = rows_y
            wb = openpyxl.load_workbook('workfile.xlsx')
            ws = wb.get_active_sheet()

            ws["%s%s" %(columns_x, rows_x)] = text
            ws["%s%s" %(columns_p, rows_p)] = page

            if count >= i:
                pass
            else:
                i = count - i2 * i
            for j in range(i):
                ws["%s%s" %(columns_y, rows_z)] = post[a]
                ws["%s%s" %(columns_z, rows_z)] = address[a]
                a += 1
                rows_z += 1

            wb.save(filename='printfile.xlsx') 

            # print data
            cmd = '"C:\Program Files\LibreOffice 5\program\soffice.exe" --headless -p printfile.xlsx'
            os.system(cmd)
            page += 1


if __name__ == "__main__":
    app = QtGui.QApplication(sys.argv)
    myapp = StartQT4()
    myapp.show()
    sys.exit(app.exec_())
=======
# -*- coding: utf-8 -*-
"""
Created on Wed Apr 27 16:50:16 2016

@author: nas-263
"""

# start program ------------------------------------------------------------

import adodbapi
import sys
from PyQt4 import QtCore, QtGui
from ui import post_text
import os
import openpyxl


class StartQT4(QtGui.QMainWindow):
    def __init__(self, parent=None):
        QtGui.QWidget.__init__(self, parent)
        self.ui = post_text.Ui_MainWindow()
        self.ui.setupUi(self)

        QtCore.QObject.connect(self.ui.pushButton, QtCore.SIGNAL("clicked()"), self.updateUi)
        #QtCore.QObject.connect(self.ui.pushButton_print, QtCore.SIGNAL("clicked()"), self.print_data)
        QtCore.QObject.connect(self.ui.pushButton_print, QtCore.SIGNAL("clicked()"), self.error)


    def error(self):
        self.ui.textEdit.append("現在、使用できません。")

    def updateUi(self):
        # input post No. or addressa
        global text
        text = (str(self.ui.lineEdit.text()))        
        self.ui.textEdit.setText("")

        # connect to the database
        database = r"./data/Database1.mdb"
        constr = 'Provider=Microsoft.Jet.OLEDB.4.0; Data Source=%s' %database
        conn = adodbapi.connect(constr)
        
        # create a cursor
        cur = conn.cursor()
        
        # extract all the data
        tablename = "24mie"
        sql = r"select 郵便番号, 県名, 市町村名, 住所 from %s" %tablename
        #sql = "select * from %s" %tablename
        cur.execute(sql)
        
        # show the cursor
        row2 = list()
        for row in cur:
            row1 = row.郵便番号[:3]+"-"+row.郵便番号[-4:]+"："+row.県名+row.市町村名+row.住所
            if row1.find(text) >= 0:
                row2 += row1,
        row2.sort()

        # close the cursor and connection
        cur.close()
        conn.close()
    
        # set data
        global post
        global address
        global count
        post = list()
        address = list()
        count = 0
        for i in row2:
            self.ui.textEdit.append(i)
            post    += i[:8],
            address += i[9:],
            count   += 1

    def print_data(self):
        wb = openpyxl.load_workbook('workfile.xlsx')
        ws = wb.get_active_sheet()
        i = 0
        alpha = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        for column in ws.columns:
            columns_count = ws.columns.index(column)
            columns_count = alpha[columns_count]
                        
            for cell in column:
                rows_count = column.index(cell) + 1

                if cell.value == "{d}":
                    rows_i = rows_count
                if cell.value == "{text}":
                    rows_x = rows_count
                    columns_x = columns_count
                if cell.value == "{post}":
                    rows_y = rows_count
                    columns_y = columns_count
                if cell.value == "{address}":
                    columns_z = columns_count
                if cell.value == "{page}":
                    rows_p = rows_count
                    columns_p = columns_count

        # write the data
        a = 0
        i = rows_i - rows_y + 1
        i2 = int((count - 1) / i) + 1

        for p in range(i2):
            page = p + 1
            rows_z = rows_y
            wb = openpyxl.load_workbook('workfile.xlsx')
            ws = wb.get_active_sheet()

            ws["%s%s" %(columns_x, rows_x)] = text
            ws["%s%s" %(columns_p, rows_p)] = page

            if count >= i:
                pass
            else:
                i = count - i2 * i
            for j in range(i):
                ws["%s%s" %(columns_y, rows_z)] = post[a]
                ws["%s%s" %(columns_z, rows_z)] = address[a]
                a += 1
                rows_z += 1

            wb.save(filename='printfile.xlsx') 

            # print data
            cmd = '"C:\Program Files\LibreOffice 5\program\soffice.exe" --headless -p printfile.xlsx'
            os.system(cmd)
            page += 1


if __name__ == "__main__":
    app = QtGui.QApplication(sys.argv)
    myapp = StartQT4()
    myapp.show()
    sys.exit(app.exec_())
>>>>>>> origin/master
